from datetime import datetime

import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from core.config import path_settings

class XLSXProcessor:

    DIVISION_NAME = "채무자"
    DIVISION_NUMBER ="사건번호"
    HEADER    = ["담당자", "채권번호", "채무자", "사건명", "사건번호", "법원", "결정일", "결정금액"]

    XLSX_FILE_NAME = path_settings.xlsx_file_name
    FONT_PATH      = path_settings.font_path
    OUTPUT_PATH    = path_settings.output_path

    
    

    def __init__(self, metadata_list,task_id):
            now = datetime.now()
            self.current_year = now.year
            self.metadata = metadata_list
            self.task_id = task_id
            self.today_str = now.strftime("%Y-%m-%d")
            pdfmetrics.registerFont(TTFont("NanumGothic", self.FONT_PATH))

            # 클래스 변수 참조 후 인스턴스 변수로 별도 저장
            self.xlsx_file_name = self.XLSX_FILE_NAME.format(year=self.current_year)
            # self.output_path = self.OUTPUT_PATH.format(today=now.strftime("%Y%m%d_%H"))
            self.output_path = self.OUTPUT_PATH.format(task_id =self.task_id)
            self.sheets = self._load_xlsx(self.xlsx_file_name)
        

            print("XLSXProceesor 초기화 완료")


    def _search_in_sheets(self, target_name: str, target_number: str, item: dict) -> bool:
        found = False
        for sheet_name, df in self.sheets.items():
            if self.DIVISION_NUMBER not in df.columns or self.DIVISION_NAME not in df.columns:
                continue
    
            matched = df[
                (df[self.DIVISION_NUMBER] == str(target_number)) &
                # (df[self.division_name] == str(target_name))
                 (df[self.DIVISION_NAME].str.contains(str(target_name), na=False))
            ]
                
            for idx in matched.index:
                found = True
                # matched.loc[idx].to_dict : 매칭된 행(row) 하나를 Series로 가져오고 dict으로변환
                # if pd.notna(v)값이 NaN인 항목 제거
                #
                row_dict = {k: v for k, v in matched.loc[idx].to_dict().items() if pd.notna(v)}
                if '채권번호' in row_dict:
                    row_dict['채권번호'] = row_dict['채권번호'].replace("-", "")
                if "강북" in sheet_name:
                    sheet_name = "강북"
                row_dict["sheet"] = sheet_name

                item["info"] = row_dict
        return found
    

    def find_number(self):
        for item in self.metadata:
            target_name = item.get('name')
            target_number = item.get('case_number')
            if target_name is None or target_number is None:
                print(f" 필수 속성 누락 - item: {item}")
                continue

            found = self._search_in_sheets(target_name, target_number, item)  # ← 현재 연도
            if(found):
                print(f"{target_name} 찾음")
            if not found:
                print("시트에서 데이터를 못찾았습니다.!!!!")
                # if str(target_number[0:4]) != str(self.year):
                #     before_year = int(self.year) -1
                #     filename = os.path.basename(new_xlsx_file)
                #     new_filename = filename.replace(str(self.year), str(before_year))
                #     new_xlsx_file = os.path.join(os.path.dirname(new_xlsx_file), new_filename)
                #     # new_xlsx_file = self.xlsx_file_name.replace(self.year, target_number[0:4])

                #     print(f"{target_name}은 {target_number[0:4]} 이므로 해당년도 파일 탐색 : {new_xlsx_file}")
                #     try:
                #         self.sheets = self._load_xlsx(new_xlsx_file)
                #         found = self._search_in_sheets(target_name, target_number, item)  # ← 다른 연도
                #         if(found):
                #             print(f"{target_name} 찾음")
                #     except FileNotFoundError:
                #         print(f"{new_xlsx_file} 파일을 찾을 수 없습니다.")
                #     finally:
                #         self.sheets =  self._load_xlsx(self.xlsx_file_name)  # 원본 복원

            if not found:
                print(f" 매치 없음 - 이름: {target_name} | 사건번호: {target_number}")
        
        # return self.metadata_list

    def create_xlsx(self):

        if os.path.exists(self.output_path):
            os.remove(self.output_path)  # 기존 파일 삭제

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        thin        = Side(style="thin", color="000000")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)
        center      = Alignment(horizontal="center", vertical="center")
        header_font = Font(name="맑은 고딕", bold=True, size=10)
        data_font   = Font(name="맑은 고딕", size=10)
        col_widths  = [8, 14, 8, 12, 16, 12, 11, 12]
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        gray_fill   = PatternFill(fill_type="solid", fgColor="D9D9D9")


        

          # ──────────────────────────────────────────
        # 1. 일반 시트들 (강북, 기타 등) → 새 양식
        # ──────────────────────────────────────────
        sheets: dict = {}
        for item in self.metadata:
            if not isinstance(item, dict):
                continue
            info       = item.get("info", {})
            sheet_name = info.get("sheet", "기타")
            sheets.setdefault(sheet_name, []).append(item)

        for sheet_name, rows in sheets.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)

                # 1행: 타이틀
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.HEADER))
                title_cell           = ws.cell(row=1, column=1, value="법원문서 전달")
                title_cell.font      = Font(name="맑은 고딕", bold=True, size=14)
                title_cell.alignment = center
                title_cell.border    = border
                ws.row_dimensions[1].height = 24

                # 2행: 헤더 (회색)
                ws.row_dimensions[2].height = 18
                for col_idx, h in enumerate(self.HEADER, start=1):
                    cell           = ws.cell(row=2, column=col_idx, value=h)
                    cell.font      = header_font
                    cell.alignment = center
                    cell.border    = border
                    cell.fill      = gray_fill

                # 열 너비
                for col_idx, width in enumerate(col_widths, start=1):
                    ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = width

            # 3행~: 데이터
            next_row = ws.max_row + 1 if ws.max_row > 2 else 3
            for i, item in enumerate(rows):
                info     = item.get("info", {})
                row_data = [
                    info.get("담당자"),
                    self._format_bond_number(info.get("채권번호")),
                    info.get("채무자"),
                    info.get("사건"),
                    info.get("사건번호"),
                    info.get("관할법원"),
                    "",
                    item.get("amount", "")
                ]
                ws.row_dimensions[next_row + i].height = 16
                for col_idx, value in enumerate(row_data, start=1):
                    cell           = ws.cell(row=next_row + i, column=col_idx, value=value)
                    cell.font      = data_font
                    cell.alignment = center
                    cell.border    = border

            # 날짜: 데이터 끝 아래 E열
            date_row            = next_row + len(rows)
            date_cell           = ws.cell(row=date_row, column=5, value=self.today_str)
            date_cell.font      = Font(name="맑은 고딕", size=10)
            date_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[date_row].height = 16

        # ──────────────────────────────────────────
        # 2. 사건데이터 시트
        # ──────────────────────────────────────────
        if "사건데이터" in wb.sheetnames:
            del wb["사건데이터"]
        ws = wb.create_sheet(title="사건데이터")

        for i, value in enumerate(self.metadata):
            row = value.get("info")
            pay = value.get("amount", "")

            if row is None:
                ws.row_dimensions[i + 1].height = 16
                continue

            sheet_name       = (row.get("sheet") or "").strip()
            should_highlight = (not pay) or (sheet_name == "개인금융")

            row_data = [
                row.get("sheet"),
                row.get("담당자"),
                self._format_bond_number(row.get("채권번호")),
                row.get("채무자"),
                row.get("사건"),
                row.get("사건번호"),
                row.get("관할법원"),
                row.get("집행권원법원"),
                row.get("집행권원사건명"),
                row.get("집행권원사건번호"),
            ]
            ws.row_dimensions[i + 1].height = 16
            for col_idx, val in enumerate(row_data, start=1):
                cell      = ws.cell(row=i + 1, column=col_idx, value=val)
                cell.font = data_font
                if should_highlight:
                    cell.fill = yellow_fill

        # ── 저장 ──
        wb.save(self.output_path)
        print(f"저장 완료: {self.output_path}")
        return self.output_path
        



    def _format_bond_number(self, value):
        """채권번호 포맷: 212056480 → 212-056480"""
        if value is None:
            return ""
        s = str(value).strip()
        if len(s) > 3:
            return f"{s[:3]}-{s[3:]}"
        return s

    def __enter__(self):
        return self
    
    def _load_xlsx(self, xlsx_file_name: str):
         return pd.read_excel(xlsx_file_name,sheet_name=None,dtype=str)


    def __exit__(self, exc_type, exc_val, exc_tb):
        pass
        #self.close()