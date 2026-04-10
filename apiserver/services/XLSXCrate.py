import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from BaseProcessor import BaseProcessor

class XLSXCreate(BaseProcessor):
    font_path = "NANUMGOTHIC.TTF"
    HEADER = ["담당자", "채권번호", "채무자", "사건명", "사건번호", "법원", "결정일", "결정금액"]
    OUTPUT_PATH = "output/result.xlsx"

    def __init__(self,metadata_file_name):
        pdfmetrics.registerFont(TTFont("NanumGothic", self.font_path))
        self.metadata = self.load_json(metadata_file_name)
        

    def create_xlsx(self):
        if os.path.exists(self.OUTPUT_PATH):
            wb = load_workbook(self.OUTPUT_PATH)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        thin   = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        header_font = Font(name="맑은 고딕", bold=True, size=10)
        data_font   = Font(name="맑은 고딕", size=10)
        col_widths  = [8, 14, 8, 12, 16, 12, 11, 12]

       # 시트별 그룹핑
       # 시트별 그룹핑
        sheets: dict = {}
        for item in self.metadata:
            info = item.get("info", {})                        # ← info 추출
            sheet_name = info.get("sheet", "기타")             # ← info에서 sheet 참조
            sheets.setdefault(sheet_name, []).append(item)

        for sheet_name, rows in sheets.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
                ws.row_dimensions[1].height = 18
                for col_idx, h in enumerate(self.HEADER, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=h)
                    cell.font      = header_font
                    cell.alignment = center
                    cell.border    = border
                for col_idx, width in enumerate(col_widths, start=1):
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

            next_row = ws.max_row + 1 if ws.max_row > 1 else 2
            for i, item in enumerate(rows):
                info = item.get("info", {})                    # ← info 추출
                row_data = [
                    info.get("담당자"),
                    self.format_bond_number(info.get("채권번호")),
                    info.get("채무자"),
                    info.get("사건"),
                    info.get("사건번호"),
                    info.get("관할법원"),
                    "",
                    "",
                ]
                ws.row_dimensions[next_row + i].height = 16
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=next_row + i, column=col_idx, value=value)
                    cell.font      = data_font
                    cell.alignment = center
                    cell.border    = border
        
        ###### 
        cur_row = 1
        ws = wb.create_sheet(title="사건데이터")
        for i, value in enumerate (self.metadata):
            row = value.get("info")
                
            if row is None:  # ← 추가
                ws.row_dimensions[cur_row + i].height = 16
                continue
            row_data = [
                row.get("sheet"),
                row.get("담당자"),
                self.format_bond_number(row.get("채권번호")),  # ← 여기만 변경
                row.get("채무자"),
                row.get("사건"),
                row.get("사건번호"),
                row.get("관할법원"),
                row.get("집행권원법원"),
                row.get("집행권원사건명"),
                row.get("집행권원사건번호"),
    
            ]
            ws.row_dimensions[cur_row + i].height = 16
            for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=cur_row + i, column=col_idx, value=value)
                    cell.font      = data_font
        wb.save(self.OUTPUT_PATH)
        print(f"저장 완료: {self.OUTPUT_PATH}")




    def format_bond_number(self, value):
        """채권번호 포맷: 212056480 → 212-056480"""
        if value is None:
            return ""
        s = str(value).strip()
        if len(s) > 3:
            return f"{s[:3]}-{s[3:]}"
        return s
    
    def run(self):
        self.create_xlsx()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass


with XLSXCreate("output/metadata.json") as reader:
    reader.run()