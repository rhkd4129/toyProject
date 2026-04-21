import os, sys
from datetime import datetime
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "/../../../")
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from apiserver.services.utils import BaseProcessor

class XLSXCreate(BaseProcessor):
    font_path   = "NANUMGOTHIC.TTF"
    HEADER      = ["담당자", "채권번호", "채무자", "사건명", "사건번호", "법원", "결정일", "결정금액"]
    

    # 변경 전
    # OUTPUT_PATH = "output/result.xlsx"

    # 변경 후

    def __init__(self, metadata_file_name):
        pdfmetrics.registerFont(TTFont("NanumGothic", self.font_path))
        self.metadata = self.load_json(metadata_file_name)
        self.OUTPUT_PATH = f"{datetime.now().strftime('%Y%m%d')}/result.xlsx"  # ← 여기로

    def create_xlsx(self):
        today_str = datetime.now().strftime("%Y-%m-%d")
        os.makedirs(os.path.dirname(self.OUTPUT_PATH), exist_ok=True)

        # ── 워크북 준비 ──
        if os.path.exists(self.OUTPUT_PATH):
            wb = load_workbook(self.OUTPUT_PATH)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # ── 공통 스타일 ──
        thin        = Side(style="thin", color="000000")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)
        center      = Alignment(horizontal="center", vertical="center")
        header_font = Font(name="맑은 고딕", bold=True, size=10)
        data_font   = Font(name="맑은 고딕", size=10)
        col_widths  = [8, 14, 8, 12, 16, 12, 11, 12]
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        gray_fill   = PatternFill(fill_type="solid", fgColor="D9D9D9")

        # ──────────────────────────────────────────
        # 1. 일반 시트들 (강북, 기타 등) → 새 양식
        # ──────────────────────────────────────────
        sheets: dict = {}
        for item in self.metadata:
            if not isinstance(item, dict):
                continue
            info       = item.get("info", {})
            sheet_name = info.get("sheet", "기타")
            sheets.setdefault(sheet_name, []).append(item)

        for sheet_name, rows in sheets.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)

                # 1행: 타이틀
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.HEADER))
                title_cell           = ws.cell(row=1, column=1, value="법원문서 전달")
                title_cell.font      = Font(name="맑은 고딕", bold=True, size=14)
                title_cell.alignment = center
                title_cell.border    = border
                ws.row_dimensions[1].height = 24

                # 2행: 헤더 (회색)
                ws.row_dimensions[2].height = 18
                for col_idx, h in enumerate(self.HEADER, start=1):
                    cell           = ws.cell(row=2, column=col_idx, value=h)
                    cell.font      = header_font
                    cell.alignment = center
                    cell.border    = border
                    cell.fill      = gray_fill

                # 열 너비
                for col_idx, width in enumerate(col_widths, start=1):
                    ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = width

            # 3행~: 데이터
            next_row = ws.max_row + 1 if ws.max_row > 2 else 3
            for i, item in enumerate(rows):
                info     = item.get("info", {})
                row_data = [
                    info.get("담당자"),
                    self.format_bond_number(info.get("채권번호")),
                    info.get("채무자"),
                    info.get("사건"),
                    info.get("사건번호"),
                    info.get("관할법원"),
                    "",
                    item.get("amount", "")
                ]
                ws.row_dimensions[next_row + i].height = 16
                for col_idx, value in enumerate(row_data, start=1):
                    cell           = ws.cell(row=next_row + i, column=col_idx, value=value)
                    cell.font      = data_font
                    cell.alignment = center
                    cell.border    = border

            # 날짜: 데이터 끝 아래 E열
            date_row             = next_row + len(rows)
            date_cell            = ws.cell(row=date_row, column=5, value=today_str)
            date_cell.font       = Font(name="맑은 고딕", size=10)
            date_cell.alignment  = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[date_row].height = 16

        # ──────────────────────────────────────────
        # 2. 사건데이터 시트 → 원래대로
        # ──────────────────────────────────────────
        ws = wb.create_sheet(title="사건데이터")
        for i, value in enumerate(self.metadata):
            row = value.get("info")
            pay = value.get("amount", "")

            if row is None:
                ws.row_dimensions[i + 1].height = 16
                continue

            sheet_name      = (row.get("sheet") or "").strip()
            should_highlight = (not pay) or (sheet_name == "개인금융")

            row_data = [
                row.get("sheet"),
                row.get("담당자"),
                self.format_bond_number(row.get("채권번호")),
                row.get("채무자"),
                row.get("사건"),
                row.get("사건번호"),
                row.get("관할법원"),
                row.get("집행권원법원"),
                row.get("집행권원사건명"),
                row.get("집행권원사건번호"),
            ]
            ws.row_dimensions[i + 1].height = 16
            for col_idx, val in enumerate(row_data, start=1):
                cell      = ws.cell(row=i + 1, column=col_idx, value=val)
                cell.font = data_font
                if should_highlight:
                    cell.fill = yellow_fill

        # ── 저장 (1번만!) ──
        wb.save(self.OUTPUT_PATH)
        print(f"저장 완료: {self.OUTPUT_PATH}")

    def format_bond_number(self, value):
        """채권번호 포맷: 212056480 → 212-056480"""
        if value is None:
            return ""
        s = str(value).strip()
        if len(s) > 3:
            return f"{s[:3]}-{s[3:]}"
        return s

    def run(self):
        self.create_xlsx()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass



# with XLSXCreate("output/metadata.json") as reader:
#     reader.run()


#     cur_row = 1
# ws = wb.create_sheet(title="사건데이터")
# write_row = cur_row  # 실제로 쓸 행 번호 별도 관리

# for value in self.metadata:  # enumerate 제거
#     row = value.get("info")
#     pay = value.get("amount", "")

#     if row is None:
#         continue

#     sheet_name = (row.get("sheet") or "").strip()

#     # 조건 충족 시 스킵
#     if (not pay) or (sheet_name == "개인금융"):
#         continue

#     row_data = [
#         row.get("sheet"),
#         row.get("담당자"),
#         self.format_bond_number(row.get("채권번호")),
#         row.get("채무자"),
#         row.get("사건"),
#         row.get("사건번호"),
#         row.get("관할법원"),
#         row.get("집행권원법원"),
#         row.get("집행권원사건명"),
#         row.get("집행권원사건번호"),
#     ]
#     ws.row_dimensions[write_row].height = 16
#     for col_idx, val in enumerate(row_data, start=1):
#         cell = ws.cell(row=write_row, column=col_idx, value=val)
#         cell.font = data_font

#     write_row += 1  # 실제로 쓴 경우에만 증가