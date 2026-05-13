import fitz
import easyocr
import numpy as np
from datetime import datetime
import re
import json
import os, sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import pandas as pd
import json
import os
class LogWrite:
    def __init__(self, filename):
        self.file = open(filename, "w", encoding="utf-8")
        self.stdout = sys.stdout

    def write(self, data):
        self.stdout.write(data)   # 콘솔에도 출력
        self.file.write(data)     # 파일에도 저장

    def flush(self):
        self.stdout.flush()
        self.file.flush()

    def close(self):
        self.file.close()
class BaseProcessor:
    """공통 기능을 모아두는 부모 클래스"""

    @staticmethod
    def load_json(path: str) -> dict | list:
        print(path)
        """JSON 파일 로드 (에러 처리 포함)"""
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f" 파일을 찾을 수 없습니다: {path}")
            sys.exit(1)
        except json.JSONDecodeError:
            print(f" JSON 파싱 실패: {path}")
            sys.exit(1)

    @staticmethod
    def save_json(path: str, data: dict | list) -> None:
        """JSON 파일 저장"""
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"저장 완료 → {path}")

    def run(self):
        raise NotImplementedError("run()을 구현해야 합니다.")
"""
    채무자별 메타데이터(이름, 사건번호, 페이지 수, 시작 페이지)만 추출해 metadata.json으로 저장하는 클래스.
    PDF 분리는 수행하지 않는다.
    """
class PDFMetadataExtractor(BaseProcessor):
    DIVISION_WORD = "채무자"
    CASE_WORD = "사건"

    def __init__(self):
        # base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.getcwd()
        # print(base_dir)
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.getcwd()


        
        pdfs = [f for f in os.listdir(base_dir) if f.lower().endswith('.pdf')]

        if len(pdfs) == 0:
            print("PDF 파일이 없습니다.")
            sys.exit(1)
        if len(pdfs) > 1:
            print(f"PDF 파일이 2개 이상입니다: {pdfs}")
            sys.exit(1)

        self.pdf_path: str = os.path.join(base_dir, pdfs[0])
        print(f"PDF 파일 감지: {self.pdf_path}")

        today = datetime.now().strftime("%Y%m%d")
        self.output_dir: str = os.path.join(base_dir, today)

        self.reader = easyocr.Reader(['ko', 'en'])
        self.doc = fitz.open(self.pdf_path)
        self.total_pages: int = len(self.doc)

        # (이름, 사건번호, [페이지 인덱스, ...]) 리스트
        self.people: list[tuple[str | None, str | None, list[int]]] = []


    def _ocr_page(self, page_index: int, fraction: int = None, section: str = "bottom") -> list:

        page = self.doc[page_index]
        pix = page.get_pixmap(dpi=300)
        img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
            pix.height, pix.width, pix.n
        )

        if fraction is None:
            # 크롭 없이 전체 페이지 OCR
            cropped = img_array
        else:
            h = img_array.shape[0]
            if section == "bottom":
                # 하단 1/fraction 크롭
                cropped = img_array[h * (fraction - 1) // fraction:, :, :]
            else:
                # 상단 1/fraction 크롭
                cropped = img_array[:h // fraction, :, :]

        return self.reader.readtext(cropped)
    



    def _extract_page_number(self, results: list) -> tuple[int | None, int | None]:
        """
        OCR 결과에서 '현재쪽/전체쪽' 패턴을 찾아 (current, total)을 반환.
        쪽번호가 없으면 (None, None)을 반환한다.
        """
        full_text = " ".join(text for (_, text, _) in results)
        match = re.search(r'(\d+)\s*/\s*(\d+)', full_text)
        if match:
            return int(match.group(1)), int(match.group(2))
        return None, None
    
    def _extract_amount(self, results: list) -> str | None:
        full_text = " ".join(text for (_, text, _) in results)
        
        # 금 6,700,014원 / 금 6,700,014 원 / 금6,700,014원 전부 커버
        match = re.search(r'[금긍]\s*([\d,]+)\s*원?', full_text)
        if match:
            return match.group(1)  # 숫자만 반환 ex) "6,700,014"
        return None
    def _extract_decision_pay(self, results):
        full_text = " ".join(text for (_, text, _) in results).replace(" ", "")
        pattern = r'[청정]구[채재]권의표시'
        return bool(re.search(pattern, full_text))  # match → search로 변경
    
    def _extract_page_number(self, results: list) -> tuple[int | None, int | None]:
        """
        OCR 결과에서 '현재쪽/전체쪽' 패턴을 찾아 (current, total)을 반환.
        쪽번호가 없으면 (None, None)을 반환한다.
        """
        full_text = " ".join(text for (_, text, _) in results)
        match = re.search(r'(\d+)\s*/\s*(\d+)', full_text)
        if match:
            return int(match.group(1)), int(match.group(2))
        return None, None
    


    def _extract_name(self, results: list) -> str | None:
        """
        OCR 결과에서 '채무자' 블록 바로 다음 블록의 텍스트를 이름으로 반환.
        찾지 못하면 None을 반환한다.
        """
        prev_clean = ""
        for (_, text, _) in results:
            clean = text.replace(" ", "")
            if prev_clean == self.DIVISION_WORD:
                name = text.strip().split()[0] if text.strip() else text.strip()
                return name[0:3]
            prev_clean = clean
        return None

    def _extract_case_number(self, results: list) -> str | None:
        """
        사건번호
        """
        prev_prev_clean = ""
        prev_clean = ""
        for (_, text, _) in results:
            clean = text.replace(" ", "")
            combined = prev_prev_clean + prev_clean
            if prev_clean == self.CASE_WORD or combined == self.CASE_WORD:
                match = re.search(r'\d{4}[가-힣]+\d+', clean)
                if match:
                    return match.group(0)
            prev_prev_clean = prev_clean
            prev_clean = clean
        return None
   

    def parse(self) -> None:
        print(f"총 {self.total_pages}페이지 감지됨\n")

        current_name: str | None = None
        current_case: str | None = None
        current_pages: list[int] = []
        current_amount: str | None = None  # 타채 금액 저장용
        i = 0

        while i < self.total_pages:
            print(f"{i + 1}페이지 OCR 읽는 중...")
            results = self._ocr_page(i)  # 전체 페이지 (쪽번호/이름/사건번호 추출용)
            current, total = self._extract_page_number(results)
            
            if current == 1:
                if current_pages:
                    self.people.append((current_name, current_case, current_pages, current_amount))

                current_name = self._extract_name(results)
                current_case = self._extract_case_number(results)
                current_amount = None  # 새 사람 시작 시 초기화
                start_index = i

                if current_case and "타채" in current_case:
                    print(f"  → 타채 사건 감지! 이름: [{current_name}] / 사건번호: [{current_case}]")
                    current_pages = [i]
                    i += 1

                    while i < start_index + total:
                        print(f"  → {i + 1}페이지 상단  크롭 OCR 읽는 중...")
                        results = self._ocr_page(i, fraction=3, section="top")  # 상단 절반만 읽기

                        if self._extract_decision_pay(results):
                            # 청구채권 문구 발견 → 금액 추출 후 내부 루프 탈출
                            current_amount = self._extract_amount(results)
                            print(f"  → 청구채권 발견! 금액: [{current_amount}]")
                            break

                        current_pages.append(i)
                        i += 1

                    # 금액 발견 여부와 관계없이 total만큼 점프
                    i = start_index + total

                else:
                    current_pages = list(range(i, i + total))
                    print(f"  → 새 사람 시작! 이름: [{current_name}] / 사건번호: [{current_case}], 총 {total}페이지")
                    print(f"  → {i + 1}~{i + total}페이지 스킵 (본문)")
                    i += total

            elif current is None:
                single_name = self._extract_name(results)
                single_case = self._extract_case_number(results)
                print(f"  → 쪽번호 없음(한 장짜리 채무자), 이름: [{single_name}] / 사건번호: [{single_case}]")
                self.people.append((single_name, single_case, [i], None))
                i += 1

            else:
                current_pages.append(i)
                print(f"  → {current}/{total} (예외처리, 포함)")
                i += 1

        if current_pages:
            self.people.append((current_name, current_case, current_pages, current_amount))

        print(f"\n총 {len(self.people)}명 감지됨\n")


    def save_metadata(self) -> None:
        """파싱 결과를 metadata.json으로만 저장한다. PDF 분리는 수행하지 않는다."""
        os.makedirs(self.output_dir, exist_ok=True)

        metadata_list = []
        for name, case_number, pages, amount in self.people:
            metadata_list.append({
                "name": name,
                "case_number": case_number,
                "pages": len(pages),
                "page_indices": pages,   # PDFSplitter가 읽을 페이지 인덱스
                "amount": amount,        # 타채 청구채권 금액 (타채 아닌 경우 None)
            })
            print(f"   이름: {name} | 사건번호: {case_number} | {len(pages)}페이지 | 금액: {amount}")

        meta_path = os.path.join(self.output_dir, "metadata.json")
        with open(meta_path, 'w', encoding='utf-8') as f:
            json.dump(metadata_list, f, ensure_ascii=False, indent=2)

        print(f"\n 메타데이터 저장 완료 → {meta_path}")
    def __enter__(self):
            return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass
    def run(self):
        self.parse()
        self.save_metadata()

# with PDFMetadataExtractor("parameter.json") as reader:
#     reader.run()

# @staticmethod란?
# @staticmethod는 클래스/인스턴스와 무관하게 동작하는 순수 함수입니다.
# 그래서 self를 파라미터로 받지 않고, 당연히 self.뭔가에도 접근할 수 없습니다.
# _load_metadata는 self.sheets, self.metadata를 수정해야 하는 함수이므로, @staticmethod가 아닌 일반 메서드로 선언해야 합니다.
class XLSXReader(BaseProcessor):
    
    division_name = "채무자"
    division_number ="사건번호"

    def __init__(self,xlsx_file_name:str,metadata_file_name:str):
        self.xlsx_file_name = xlsx_file_name
        self.metadata_file_name = metadata_file_name

        self.metadata = self.load_json(metadata_file_name)
        self.sheets = self._load_xlsx(xlsx_file_name)      

        self.year = xlsx_file_name[0:4]


        
        # 드모르간 법칙으로 이해하면 쉽습니다:
        # ```
        # "A도 있고 B도 있어야 한다"
        # = "A가 없거나 B가 없으면 skip"
        # = not in A or not in B
    def _search_in_sheets(self, target_name: str, target_number: str, item: dict) -> bool:
        found = False
        for sheet_name, df in self.sheets.items():
            if self.division_number not in df.columns or self.division_name not in df.columns:
                continue
    
            matched = df[
                (df[self.division_number] == str(target_number)) &
                # (df[self.division_name] == str(target_name))
                 (df[self.division_name].str.contains(str(target_name), na=False))
            ]
                
            for idx in matched.index:
                found = True
                # matched.loc[idx].to_dict : 매칭된 행(row) 하나를 Series로 가져오고 dict으로변환
                # if pd.notna(v)값이 NaN인 항목 제거
                #
                row_dict = {k: v for k, v in matched.loc[idx].to_dict().items() if pd.notna(v)}
                if '채권번호' in row_dict:
                    row_dict['채권번호'] = row_dict['채권번호'].replace("-", "")
                if "강북" in sheet_name:
                    sheet_name = "강북"
                row_dict["sheet"] = sheet_name

                item["info"] = row_dict
        return found


    def find_number(self):
        for item in self.metadata:
            target_name = item.get('name')
            target_number = item.get('case_number')
            if target_name is None or target_number is None:
                print(f" 필수 속성 누락 - item: {item}")
                continue

            found = self._search_in_sheets(target_name, target_number, item)  # ← 현재 연도
            if(found):
                print(f"{target_name} 찾음")
            if not found:
                print("시트에서 데이터를 못찾았습니다.!!!!")
                if str(target_number[0:4]) != str(self.year):
                    print(f"{self.target_name}은{self.year}년 사건이 아닙니다 pass")
                else:
                    print(f" 매치 없음 - 이름: {target_name} | 사건번호: {target_number}")
                    
                #     before_year = int(self.year) -1
                #     filename = os.path.basename(new_xlsx_file)
                #     new_filename = filename.replace(str(self.year), str(before_year))
                #     new_xlsx_file = os.path.join(os.path.dirname(new_xlsx_file), new_filename)
                #     # new_xlsx_file = self.xlsx_file_name.replace(self.year, target_number[0:4])

                #     print(f"{target_name}은 {target_number[0:4]} 이므로 해당년도 파일 탐색 : {new_xlsx_file}")
                #     try:
                #         self.sheets = self._load_xlsx(new_xlsx_file)
                #         found = self._search_in_sheets(target_name, target_number, item)  # ← 다른 연도
                #         if(found):
                #             print(f"{target_name} 찾음")
                #     except FileNotFoundError:
                #         print(f"{new_xlsx_file} 파일을 찾을 수 없습니다.")
                #     finally:
                #         self.sheets =  self._load_xlsx(self.xlsx_file_name)  # 원본 복원

            # if not found:
            #     print(f" 매치 없음 - 이름: {target_name} | 사건번호: {target_number}")

    def _load_xlsx(self, xlsx_file_name: str):
         return pd.read_excel(xlsx_file_name,sheet_name=None,dtype=str)



    def run(self):
        # self._load_metadata(self.xlsx_file,self.metadata_file)
        self.find_number()
        with open(self.metadata_file_name, "w", encoding="utf-8") as f:
                json.dump(self.metadata, f, ensure_ascii=False, indent=2)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass
        #self.close()



# with XLSXReader("2026사건부.xlsx","output/metadata.json") as reader:
#     reader.run()

class XLSXCreate(BaseProcessor):
    font_path = "NANUMGOTHIC.TTF"
    HEADER    = ["담당자", "채권번호", "채무자", "사건명", "사건번호", "법원", "결정일", "결정금액"]

    def __init__(self, metadata_file_name):
        pdfmetrics.registerFont(TTFont("NanumGothic", self.font_path))
        self.metadata = self.load_json(metadata_file_name)

        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.getcwd()

        now              = datetime.now()
        today            = now.strftime('%Y%m%d')
        self.today_str   = now.strftime("%Y-%m-%d")
        self.OUTPUT_PATH = os.path.join(base_dir, today, "result.xlsx")
        os.makedirs(os.path.dirname(self.OUTPUT_PATH), exist_ok=True)

    def create_xlsx(self):
        # ── 워크북 준비 ──
        if os.path.exists(self.OUTPUT_PATH):
            wb = load_workbook(self.OUTPUT_PATH)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # ── 공통 스타일 ──
        thin        = Side(style="thin", color="000000")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)
        center      = Alignment(horizontal="center", vertical="center")
        header_font = Font(name="맑은 고딕", bold=True, size=10)
        data_font   = Font(name="맑은 고딕", size=10)
        col_widths  = [8, 14, 8, 12, 16, 12, 11, 12]
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        gray_fill   = PatternFill(fill_type="solid", fgColor="D9D9D9")

        # ──────────────────────────────────────────
        # 1. 일반 시트들 (강북, 기타 등) → 새 양식
        # ──────────────────────────────────────────
        sheets: dict = {}
        for item in self.metadata:
            if not isinstance(item, dict):
                continue
            info       = item.get("info", {})
            sheet_name = info.get("sheet", "기타")
            sheets.setdefault(sheet_name, []).append(item)

        for sheet_name, rows in sheets.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)

                # 1행: 타이틀
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.HEADER))
                title_cell           = ws.cell(row=1, column=1, value="법원문서 전달")
                title_cell.font      = Font(name="맑은 고딕", bold=True, size=14)
                title_cell.alignment = center
                title_cell.border    = border
                ws.row_dimensions[1].height = 24

                # 2행: 헤더 (회색)
                ws.row_dimensions[2].height = 18
                for col_idx, h in enumerate(self.HEADER, start=1):
                    cell           = ws.cell(row=2, column=col_idx, value=h)
                    cell.font      = header_font
                    cell.alignment = center
                    cell.border    = border
                    cell.fill      = gray_fill

                # 열 너비
                for col_idx, width in enumerate(col_widths, start=1):
                    ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = width

            # 3행~: 데이터
            next_row = ws.max_row + 1 if ws.max_row > 2 else 3
            for i, item in enumerate(rows):
                info     = item.get("info", {})
                row_data = [
                    info.get("담당자"),
                    self.format_bond_number(info.get("채권번호")),
                    info.get("채무자"),
                    info.get("사건"),
                    info.get("사건번호"),
                    info.get("관할법원"),
                    "",
                    item.get("amount", "")
                ]
                ws.row_dimensions[next_row + i].height = 16
                for col_idx, value in enumerate(row_data, start=1):
                    cell           = ws.cell(row=next_row + i, column=col_idx, value=value)
                    cell.font      = data_font
                    cell.alignment = center
                    cell.border    = border

            # 날짜: 데이터 끝 아래 E열
            date_row            = next_row + len(rows)
            date_cell           = ws.cell(row=date_row, column=5, value=self.today_str)
            date_cell.font      = Font(name="맑은 고딕", size=10)
            date_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[date_row].height = 16

        # ──────────────────────────────────────────
        # 2. 사건데이터 시트
        # ──────────────────────────────────────────
        if "사건데이터" in wb.sheetnames:
            del wb["사건데이터"]
        ws = wb.create_sheet(title="사건데이터")

        for i, value in enumerate(self.metadata):
            row = value.get("info")
            pay = value.get("amount", "")

            if row is None:
                ws.row_dimensions[i + 1].height = 16
                continue

            sheet_name       = (row.get("sheet") or "").strip()
            should_highlight = (not pay) or (sheet_name == "개인금융")

            row_data = [
                row.get("sheet"),
                row.get("담당자"),
                self.format_bond_number(row.get("채권번호")),
                row.get("채무자"),
                row.get("사건"),
                row.get("사건번호"),
                row.get("관할법원"),
                row.get("집행권원법원"),
                row.get("집행권원사건명"),
                row.get("집행권원사건번호"),
            ]
            ws.row_dimensions[i + 1].height = 16
            for col_idx, val in enumerate(row_data, start=1):
                cell      = ws.cell(row=i + 1, column=col_idx, value=val)
                cell.font = data_font
                if should_highlight:
                    cell.fill = yellow_fill

        # ── 저장 ──
        wb.save(self.OUTPUT_PATH)
        print(f"저장 완료: {self.OUTPUT_PATH}")

    def run(self):
        self.create_xlsx()

    def format_bond_number(self, value):
        """채권번호 포맷: 212056480 → 212-056480"""
        if value is None:
            return ""
        s = str(value).strip()
        if len(s) > 3:
            return f"{s[:3]}-{s[3:]}"
        return s

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass



import traceback
if __name__ == "__main__":
    base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.getcwd()
    today = datetime.now().strftime("%Y%m%d")

    filename = datetime.now().strftime("log_%Y%m%d_%H%M%S.txt")
    log = LogWrite(filename)
    sys.stdout = log

    start_time = datetime.now()
    print(f"▶ 실행 시작: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 30)

    xlsx_file    = os.path.join(base_dir, "2026사건부.xlsx")
    metadata_file = os.path.join(base_dir, today, "metadata.json")
    os.makedirs(os.path.join(base_dir, today), exist_ok=True)

    try:
        print("1. metadata.json 작성")
        with PDFMetadataExtractor() as extractor:
            extractor.run()

        print("2. 사건부 조회 후 매칭")
        with XLSXReader(xlsx_file, metadata_file) as reader:
            reader.run()

        print("3. 시트 생성")
        with XLSXCreate(metadata_file) as creator:
            creator.run()

    except Exception:
        traceback.print_exc()

    finally:
        end_time = datetime.now()
        elapsed = end_time - start_time
        print("=" * 30)
        print(f"실행 종료: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"총 소요시간: {elapsed.seconds}초 ({elapsed})")
        sys.stdout = log.stdout
        log.close()
        input("\n종료하려면 Enter를 누르세요...")


# pyinstaller --onefile --add-data "NANUMGOTHIC.TTF;." --paths "../../../" Executor.py